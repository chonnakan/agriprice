# -*- coding: utf-8 -*-
import requests
import json
import openpyxl
import urllib
import os
import xlrd

master_key = 'AIzaSyA9RDsFBy2tE2PXnx9ecqCxN7mBXMsuHCE'

key_list = ['AIzaSyAEGQvWQy2rm4NQje7Ez1ak6WMZUYPhM78', 'AIzaSyA4Hq8KF595FhBeavKDkkHPtnJZG9cCbvc',
            'AIzaSyCPDaFCE7t641orftenf4qy0rFwtzrpe24', 'AIzaSyAsZTiQ8czDsW_nfd3VWfP021GQCn-YXj0',
            'AIzaSyC79xe3OhOLTW07_iHnY7nzZ4DZu_p8yyA']

key_words = ['atm', '']


def nearby_search(lat, lng, name):
    url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location='
    url = url + lat + ',' + lng + '&keyword=' + urllib.parse.quote(name.encode('utf8')) + '&rankby=distance&key=' + master_key
    req = requests.get(url)
    res = json.loads(req.content)
    list = []
    #print url
    err_message = ''
    if res['status'] == 'OK':
        #print url
        for p in res['results']:
            list.append(p)
        #time.sleep(2)
        while 'next_page_token' in res:
            page_token = res['next_page_token']
            url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location='
            url = url + lat + ',' + lng + '&keyword=' + urllib.parse.quote(name.encode('utf8')) + '&rankby=distance&key=' + master_key + '&pagetoken=' + page_token
            #print url
            req = requests.get(url)
            res = json.loads(req.content)
            if res['status'] == 'OK':
                for p in res['results']:
                    list.append(p)
            else:
                while res['status'] != 'OK':
                    url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location='
                    url = url + lat + ',' + lng + '&keyword=' + urllib.parse.quote(name.encode('utf8')) + '&rankby=distance&key=' + master_key + '&pagetoken=' + page_token
                    #print url
                    req = requests.get(url)
                    res = json.loads(req.content)
                    if res['status'] == 'OK':
                        for p in res['results']:
                            list.append(p)
            #time.sleep(2)
        return list, 'OK', err_message
    elif res['status'] == 'ZERO_RESULTS':
        if 'error_message' in res:
            err_message = res['error_message']
        return [], 'ZERO_RESULTS', err_message
    elif res['status'] == 'OVER_QUERY_LIMIT':
        if 'error_message' in res:
            err_message = res['error_message']
        return [], 'OVER_QUERY_LIMIT', err_message
    elif res['status'] == 'REQUEST_DENIED':
        if 'error_message' in res:
            err_message = res['error_message']
        return [], 'REQUEST_DENIED', err_message
    elif res['status'] == 'INVALID_REQUEST':
        if 'error_message' in res:
            err_message = res['error_message']
        return [], 'INVALID_REQUEST', err_message
    else:
        if 'error_message' in res:
            err_message = res['error_message']
        return [], res['status'], err_message


def get_direction(lat1, lng1, lat2, lng2, k):
    url = 'https://maps.googleapis.com/maps/api/directions/json?origin='
    url = url + lat1 + ',' + lng1 + '&destination=' + lat2 + ',' + lng2 + '&avoid=tolls&departure_time=1514764799000&key=' + k
    req = requests.get(url)
    res = json.loads(req.content)
    err_message = ''
    d1 = d2 = d3 = 0
    if res['status'] == 'OK':
        d1 = res['routes'][0]['legs'][0]['distance']['value']
        d2 = res['routes'][0]['legs'][0]['duration']['value']
        d3 = res['routes'][0]['legs'][0]['duration_in_traffic']['value']
        return [d1, d2, d3], 'OK', err_message
    elif res['status'] == 'ZERO_RESULTS':
        if 'error_message' in res:
            err_message = res['error_message']
        return [d1, d2, d3], 'ZERO_RESULTS', err_message
    elif res['status'] == 'OVER_QUERY_LIMIT':
        if 'error_message' in res:
            err_message = res['error_message']
        return [d1, d2, d3], 'OVER_QUERY_LIMIT', err_message
    elif res['status'] == 'REQUEST_DENIED':
        if 'error_message' in res:
            err_message = res['error_message']
        return [d1, d2, d3], 'REQUEST_DENIED', err_message
    elif res['status'] == 'INVALID_REQUEST':
        if 'error_message' in res:
            err_message = res['error_message']
        return [d1, d2, d3], 'INVALID_REQUEST', err_message
    else:
        if 'error_message' in res:
            err_message = res['error_message']
        return [d1, d2, d3], res['status'], err_message


def get_nearest_dist(lat, lng, keyword, key):

    places, status, message = nearby_search(lat, lng, keyword)
    #time.sleep(0.1)
    # print status,
    d = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    n = ['', '', '']
    p1 = None
    p2 = None
    p3 = None

    for i in range(0, len(places)):
        places[i]['name'] = places[i]['name'].lower()
        #print places[i]['name']
        if u'atm' in places[i]['name'] or u'เอทีเอ็ม' in places[i]['name']:
            n[0] = places[i]['name']
            p1 = places[i]
            dd, status, message = get_direction(lat, lng, str(p1['geometry']['location']['lat']),
                                               str(p1['geometry']['location']['lng']), key)
            d[0] = dd[0]
            d[1] = dd[1]
            d[2] = dd[2]
            break
    #print '1-----------' + n[0]
    #time.sleep(0.1)
    for i in range(0, len(places)):
        places[i]['name'] = places[i]['name'].lower()
        #print places[i]['name']
        if u'atm' in places[i]['name'] or u'เอทีเอ็ม' in places[i]['name']:
            if u'ธ.ก.ส.' in places[i]['name'] or u'ธกส' in places[i]['name'] or u'เพื่อการเกษตร' in places[i]['name'] or \
                            u'baac' in places[i]['name'] or u'agriculture' in places[i]['name'] or \
                            u'ออมสิน' in places[i]['name'] or u'gsb' in places[i]['name'] or \
                            u'government savning' in places[i]['name'] or u'อิสลาม' in places[i]['name'] or \
                            u'ibank' in places[i]['name'] or  u'islam' in places[i]['name']:
                n[1] = places[i]['name']
                p2 = places[i]
                break
    #print '2-----------' + n[1]

    if n[1] != '':
        dd, status, message = get_direction(lat, lng, str(p2['geometry']['location']['lat']),
                                            str(p2['geometry']['location']['lng']), key)
        d[3] = dd[0]
        d[4] = dd[1]
        d[5] = dd[2]
    #time.sleep(0.1)
    for i in range(0, len(places)):
        places[i]['name'] = places[i]['name'].lower()
        #print places[i]['name']
        if u'atm' in places[i]['name'] or u'เอทีเอ็ม' in places[i]['name']:
            if not(u'ธ.ก.ส.' in places[i]['name'] or u'ธกส' in places[i]['name'] or u'เพื่อการเกษตร' in places[i]['name'] or u'baac' in places[i]['name'] or
                           u'agriculture' in places[i]['name'] or u'ออมสิน' in places[i]['name'] or u'gsb' in places[i]['name'] or u'government savning' in places[i]['name'] or
                           u'อิสลาม' in places[i]['name'] or u'ibank' in places[i]['name'] or u'islam' in places[i]['name']):
                n[2] = places[i]['name']
                p3 = places[i]
                break
    #print '3-----------' + n[2]
    if n[2] != '':
        dd, status, message = get_direction(lat, lng, str(p3['geometry']['location']['lat']),
                                            str(p3['geometry']['location']['lng']), key)
        d[6] = dd[0]
        d[7] = dd[1]
        d[8] = dd[2]

    return d, n, status, message


def main():
    wb = xlrd.open_workbook('village_stat_2.xlsx')
    ws = wb.sheet_by_index(0)
    k_i = 0
    xlname = 'check_all_atm.xlsx'

    if os.path.exists(os.path.join(os.path.dirname(os.path.abspath(__file__)), xlname)):
        owb = openpyxl.load_workbook(xlname, read_only=False)
    else:
        owb = openpyxl.Workbook()

    if len(owb.sheetnames) == 0:
        ows = owb.create_sheet('Sheet')
    else:
        ows = owb.active

    for i in range(1, ws.nrows):
        if k_i == 5:
            break
        print('%d: %f' % (k_i, float(i)/ws.nrows)),
        print(ws.cell_value(i, 0)),
        print(ws.cell_value(i, 1)),
        status = ''
        #print float(ws.cell_value(i, 2)), float(ws.cell_value(i, 3)),
        #lng, lat = myCood(float(ws.cell_value(i, 2)), float(ws.cell_value(i, 3)), inverse=True)
        #print lat, lng,
        lat = str(ws.cell_value(i, 93))
        lng = str(ws.cell_value(i, 92))
        print (lat, lng),
        d, names, status, message = get_nearest_dist(lat, lng, u'atm', master_key)
        print(status),

        while status == 'OVER_QUERY_LIMIT' and k_i < 5:
            k_i += 1
            d, names, status, message = get_nearest_dist(lat, lng, u'atm', master_key)
            print(status),

        if status == 'OK':
            for j in range(0, 4):
                c = ows.cell(row=i, column=j+1, value=ws.cell_value(i, j))
            c = ows.cell(row=i, column=5, value=names[0])
            c = ows.cell(row=i, column=6, value=d[0])
            c = ows.cell(row=i, column=7, value=d[1])
            c = ows.cell(row=i, column=8, value=d[2])
            c = ows.cell(row=i, column=9, value=names[1])
            c = ows.cell(row=i, column=10, value=d[3])
            c = ows.cell(row=i, column=11, value=d[4])
            c = ows.cell(row=i, column=12, value=d[5])
            c = ows.cell(row=i, column=13, value=names[2])
            c = ows.cell(row=i, column=14, value=d[6])
            c = ows.cell(row=i, column=15, value=d[7])
            c = ows.cell(row=i, column=16, value=d[8])
            owb.save(xlname)
        print('')

if __name__ == "__main__":
    main()
